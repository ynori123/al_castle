from openpyxl import load_workbook
from src.model import (
    Castle,
    Restaurant
)
from src.database import get_db



def load_data() -> tuple[list[Castle], list[Restaurant]]:
    castle_data = []
    try:
        with load_workbook("castle_data.xlsx") as wb:
            # 飲食店データの読み込み
            ws = wb["castle_data"]
            for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                if col[0].value.isdecimal() == False:
                    # 1列目が数字でない場合はスキップ
                    continue
                castle = Castle(
                    name=col[1].value,
                    prefecture=col[2].value,
                    lat=float(col[3].value),
                    lng=float(col[4].value),
                    holiday=col[5].value,
                    admission_time=col[6].value,
                    admission_fee=col[7].value,
                    stamp=col[8].value
                )
                castle_data.append(castle)
            restaurant_data = []
            ws = wb["meal"]
            for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                if col[0].value.isdecimal() == False:
                    # 1列目が数字でない場合はスキップ
                    continue
                restaurant = Restaurant(
                    castle_id=int(col[1].value),
                    name=col[2].value,
                    time=col[3].value,
                    holiday=col[5].value,
                    genre=col[6].value,
                    url=col[7].value,
                )
                restaurant_data.append(castle)
            return castle_data, restaurant_data
    except Exception as e:
        print(e)
        return 

def write_data(castle_data: list[Castle], restaurant_data: list[Restaurant]):
    with get_db() as db:
        db.add_all(castle_data)
        db.add_all(restaurant_data)
        db.commit()
    return
